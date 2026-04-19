#!/usr/bin/env python3
"""expense-tracker Excel operations."""

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DETAIL_SHEET = "明细"
SUMMARY_SHEET = "汇总"

# 列顺序：金额移至备注左侧
HEADERS    = ["序号", "费用类型", "发生日期", "事由/摘要", "付款方式", "发票状态", "经办人", "金额（元）", "备注", "录入时间"]
COL_WIDTHS = [6,      12,         12,          32,           10,          10,          10,      14,           22,     18]

# 列索引（1-based）
COL_SEQ, COL_TYPE, COL_DATE, COL_DESC, COL_PAY, COL_INV, COL_PERSON, COL_AMT, COL_NOTE, COL_TIME = range(1, 11)

# 报账系统
EXPENSE_TYPES_REIMBURSEMENT = [
    "办公用品", "邮电费", "版面费", "印刷费", "专利费",
    "设备燃动费", "维修费", "委托业务费", "水电费", "租赁费",
    "市内交通费", "活动费", "培训费", "会议费", "加班餐费",
    "公务招待费", "公务用车费", "因公出国费",
    "国内差旅费", "国际差旅费", "其他日常费用",
]
# 酬金系统
EXPENSE_TYPES_REMUNERATION = [
    "编审费", "答辩费", "奖励", "课酬", "评审费",
    "科研劳务费", "人才津贴", "咨询费", "稿酬",
    "科研绩效", "中学生奖助学金", "外籍专家劳务费",
    "高层次人才免税", "考务费", "其他劳务",
]
EXPENSE_TYPES = EXPENSE_TYPES_REIMBURSEMENT + EXPENSE_TYPES_REMUNERATION
NO_INVOICE_TYPES = set(EXPENSE_TYPES_REMUNERATION)

C_DARK_BLUE, C_MID_BLUE, C_LIGHT_BLUE = "1F4E79", "2E75B6", "D6E4F0"
C_ALT, C_WHITE, C_GOLD, C_SUBTOTAL = "EBF3FB", "FFFFFF", "FFF2CC", "E2EFDA"

THIN   = Side(style="thin",   color="CCCCCC")
MEDIUM = Side(style="medium", color="999999")
BORDER        = Border(left=THIN,   right=THIN,   top=THIN,   bottom=THIN)
BORDER_BOTTOM = Border(left=THIN,   right=THIN,   top=THIN,   bottom=MEDIUM)

AMOUNT_FMT = '#,##0.00'

FONT_BODY     = Font(size=10, name="微软雅黑")
FONT_HEADER   = Font(bold=True, color=C_WHITE, size=11, name="微软雅黑")
FONT_TITLE    = Font(bold=True, color=C_WHITE, size=15, name="微软雅黑")
FONT_TOTAL    = Font(bold=True, size=11, name="微软雅黑", color=C_DARK_BLUE)
FONT_SUBTOTAL = Font(bold=True, size=10, name="微软雅黑", color="375623")
FONT_LABEL    = Font(bold=True, size=11, name="微软雅黑")

FILL_HEADER   = PatternFill("solid", start_color=C_DARK_BLUE)
FILL_TITLE    = PatternFill("solid", start_color=C_MID_BLUE)
FILL_ALT      = PatternFill("solid", start_color=C_ALT)
FILL_WHITE    = PatternFill("solid", start_color=C_WHITE)
FILL_GOLD     = PatternFill("solid", start_color=C_GOLD)
FILL_SUBTOTAL = PatternFill("solid", start_color=C_SUBTOTAL)

TITLE_ROW, HEADER_ROW, DATA_START = 1, 2, 3
NUM_COLS = len(HEADERS)


def _setup_detail_sheet(ws, title: str):
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=NUM_COLS)
    tc = ws.cell(row=TITLE_ROW, column=1, value=title)
    tc.font = FONT_TITLE
    tc.fill = FILL_TITLE
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 36

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HEADER_ROW].height = 22
    ws.freeze_panes = f"A{DATA_START}"


def _style_data_cell(c, fill, font=None, align=None, border=BORDER, fmt=None):
    c.fill = fill
    c.font = font or FONT_BODY
    c.alignment = align or Alignment(vertical="center")
    c.border = border
    if fmt:
        c.number_format = fmt


def _write_subtotal_row(ws, row: int, label: str, amt_col_letter: str, start_row: int, end_row: int):
    """Insert a subtotal row; seq cell left empty so numbering is uninterrupted."""
    sc = ws.cell(row=row, column=COL_SEQ, value="")
    sc.fill = FILL_SUBTOTAL
    sc.border = BORDER

    ws.merge_cells(start_row=row, start_column=COL_TYPE, end_row=row, end_column=COL_PERSON)
    lc = ws.cell(row=row, column=COL_TYPE, value=f"{label} 小计")
    lc.font = FONT_SUBTOTAL
    lc.fill = FILL_SUBTOTAL
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = BORDER

    ac = ws.cell(row=row, column=COL_AMT,
                 value=f"=SUM({get_column_letter(COL_AMT)}{start_row}:{get_column_letter(COL_AMT)}{end_row})")
    ac.font = FONT_SUBTOTAL
    ac.fill = FILL_SUBTOTAL
    ac.alignment = Alignment(vertical="center", horizontal="center")
    ac.border = BORDER
    ac.number_format = AMOUNT_FMT

    for col in (COL_NOTE, COL_TIME):
        c = ws.cell(row=row, column=col, value="")
        c.fill = FILL_SUBTOTAL
        c.border = BORDER
    ws.row_dimensions[row].height = 18


def _write_total_row(ws, row: int):
    """Grand total row; seq cell left empty."""
    sc = ws.cell(row=row, column=COL_SEQ, value="")
    sc.fill = FILL_GOLD
    sc.border = BORDER

    ws.merge_cells(start_row=row, start_column=COL_TYPE, end_row=row, end_column=COL_PERSON)
    lc = ws.cell(row=row, column=COL_TYPE, value="合  计")
    lc.font = FONT_TOTAL
    lc.fill = FILL_GOLD
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = BORDER

    ac = ws.cell(row=row, column=COL_AMT,
                 value=f"=SUMIF({get_column_letter(COL_AMT)}{DATA_START}:{get_column_letter(COL_AMT)}{row-1},\">0\")")
    ac.font = FONT_TOTAL
    ac.fill = FILL_GOLD
    ac.alignment = Alignment(vertical="center", horizontal="center")
    ac.border = BORDER
    ac.number_format = AMOUNT_FMT

    for col in (COL_NOTE, COL_TIME):
        c = ws.cell(row=row, column=col, value="")
        c.fill = FILL_GOLD
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def _build_summary_sheet(ws, title: str):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value=f"{title} — 汇总")
    t.font = FONT_TITLE
    t.fill = FILL_TITLE
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    d = f"'{DETAIL_SHEET}'"
    # Note: after restructure, type=col B(2), amount=col D(4), pay=col F(6), inv=col G(7)
    sections = [
        ("总览", [
            ("总金额（元）", f"=SUMIF({d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)},\">0\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})"),
            ("总条数",       f"=COUNTA({d}!A:A)-2"),
        ]),
        ("按付款方式", [
            ("对公合计", f"=SUMIF({d}!{get_column_letter(COL_PAY)}:{get_column_letter(COL_PAY)},\"对公\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})"),
            ("对私合计", f"=SUMIF({d}!{get_column_letter(COL_PAY)}:{get_column_letter(COL_PAY)},\"对私\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})"),
        ]),
        ("按发票状态", [
            ("有发票合计", f"=SUMIF({d}!{get_column_letter(COL_INV)}:{get_column_letter(COL_INV)},\"有发票\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})"),
            ("无发票合计", f"=SUMIF({d}!{get_column_letter(COL_INV)}:{get_column_letter(COL_INV)},\"无发票\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})"),
            ("待补合计",   f"=SUMIF({d}!{get_column_letter(COL_INV)}:{get_column_letter(COL_INV)},\"待补\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})"),
        ]),
        ("按费用类型", [
            (et, f"=SUMIF({d}!{get_column_letter(COL_TYPE)}:{get_column_letter(COL_TYPE)},\"{et}\",{d}!{get_column_letter(COL_AMT)}:{get_column_letter(COL_AMT)})")
            for et in EXPENSE_TYPES
        ]),
    ]

    row = 2
    for section_name, items in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sh = ws.cell(row=row, column=1, value=f"  {section_name}")
        sh.font = Font(bold=True, color=C_WHITE, size=11, name="微软雅黑")
        sh.fill = PatternFill("solid", start_color=C_MID_BLUE)
        sh.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        for label, formula in items:
            a = ws.cell(row=row, column=1, value=label)
            b = ws.cell(row=row, column=2, value=formula)
            fill = FILL_GOLD if label == "总金额（元）" else (FILL_ALT if row % 2 == 0 else FILL_WHITE)
            for c in (a, b):
                c.fill = fill
                c.border = BORDER
                c.alignment = Alignment(vertical="center", horizontal="right" if c.column == 2 else "left")
            a.font = FONT_TOTAL if label == "总金额（元）" else FONT_LABEL
            b.font = FONT_TOTAL if label == "总金额（元）" else FONT_BODY
            b.number_format = AMOUNT_FMT
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1


def _create_workbook(path: Path, title: str):
    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = DETAIL_SHEET
    _setup_detail_sheet(ws_detail, title)
    ws_sum = wb.create_sheet(SUMMARY_SHEET)
    _build_summary_sheet(ws_sum, title)
    wb.save(path)


def _rebuild_detail(ws, all_records: list, now: str):
    """Rewrite all data rows with merged type cells, subtotals, and grand total."""
    # Clear existing data rows (keep title row 1 and header row 2)
    if ws.max_row >= DATA_START:
        ws.delete_rows(DATA_START, ws.max_row - DATA_START + 1)

    # Group records by expense type, preserving insertion order
    groups: dict[str, list] = {}
    for rec in all_records:
        t = rec.get("费用类型", "其他")
        groups.setdefault(t, []).append(rec)

    row = DATA_START
    seq = 1
    for expense_type, recs in groups.items():
        group_start = row
        fill = FILL_ALT if (seq % 2 == 0) else FILL_WHITE

        for i, rec in enumerate(recs):
            invoice = "无发票" if expense_type in NO_INVOICE_TYPES else rec.get("发票状态", "待补")
            values = {
                COL_SEQ:    seq,
                COL_TYPE:   expense_type if i == 0 else None,
                COL_DATE:   rec.get("发生日期", "待补"),
                COL_DESC:   rec.get("事由", "待补"),
                COL_PAY:    rec.get("付款方式", "待补"),
                COL_INV:    invoice,
                COL_PERSON: rec.get("经办人", ""),
                COL_AMT:    rec.get("金额", "待补"),
                COL_NOTE:   rec.get("备注", ""),
                COL_TIME:   rec.get("录入时间", now),
            }
            for col, val in values.items():
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill
                c.border = BORDER
                c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=(col == COL_DESC))
                if col == COL_AMT and isinstance(val, (int, float)):
                    c.number_format = AMOUNT_FMT
                    c.font = Font(bold=True, size=10, name="微软雅黑")
                else:
                    c.font = FONT_BODY
            ws.row_dimensions[row].height = 18
            seq += 1
            row += 1

        # Merge type column for this group
        if len(recs) > 1:
            ws.merge_cells(start_row=group_start, start_column=COL_TYPE,
                           end_row=row - 1,       end_column=COL_TYPE)
            tc = ws.cell(row=group_start, column=COL_TYPE)
            tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Subtotal row for this group
        _write_subtotal_row(ws, row, expense_type, get_column_letter(COL_AMT), group_start, row - 1)
        row += 1

    # Grand total
    _write_total_row(ws, row)


def append_rows(file_path: str, title: str, records: list) -> dict:
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Tag each new record with current time
    for rec in records:
        rec.setdefault("录入时间", now)
        expense_type = rec.get("费用类型", "其他")
        if expense_type in NO_INVOICE_TYPES:
            rec["发票状态"] = "无发票"

    if not path.exists():
        _create_workbook(path, title)
        wb = load_workbook(path)
        ws = wb[DETAIL_SHEET]
        all_records = records
    else:
        wb = load_workbook(path)
        ws = wb[DETAIL_SHEET] if DETAIL_SHEET in wb.sheetnames else wb.active
        # Read existing data rows (skip title row 1, header row 2, subtotal/total rows)
        existing = []
        for r in range(DATA_START, ws.max_row + 1):
            seq_val = ws.cell(row=r, column=COL_SEQ).value
            if not isinstance(seq_val, int):
                continue  # skip subtotal/total rows
            existing.append({
                "费用类型":  ws.cell(row=r, column=COL_TYPE).value or "",
                "发生日期":  ws.cell(row=r, column=COL_DATE).value or "待补",
                "事由":      ws.cell(row=r, column=COL_DESC).value or "待补",
                "付款方式":  ws.cell(row=r, column=COL_PAY).value or "待补",
                "发票状态":  ws.cell(row=r, column=COL_INV).value or "待补",
                "经办人":    ws.cell(row=r, column=COL_PERSON).value or "",
                "金额":      ws.cell(row=r, column=COL_AMT).value,
                "备注":      ws.cell(row=r, column=COL_NOTE).value or "",
                "录入时间":  ws.cell(row=r, column=COL_TIME).value or "",
            })
        all_records = existing + records

    _rebuild_detail(ws, all_records, now)
    wb.save(path)

    data_count = sum(1 for rec in all_records if True)
    return {"added": len(records), "total_rows": data_count, "file": str(path)}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--title", default="报账明细")
    parser.add_argument("--data", required=True)
    args = parser.parse_args()

    records = json.loads(args.data)
    result = append_rows(args.file, args.title, records)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
